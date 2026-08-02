import pandas as pd
import os
import re
from datetime import datetime
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def clean_rating(r):
    if pd.isna(r) or r == "": return None
    r = str(r).replace(',', '.').strip().replace('"','')
    try:
        return float(r)
    except:
        return None

def clean_website(url):
    if pd.isna(url) or not str(url).strip():
        return ""
    url = str(url).strip()
    parsed = urlparse(url)
    if parsed.scheme and parsed.netloc:
        clean = f"{parsed.scheme}://{parsed.netloc}{parsed.path}".rstrip('/')
        return clean
    return url.split('?')[0]

def format_phone(p):
    if pd.isna(p) or not str(p).strip():
        return ""
    p = str(p).strip()
    digits = re.sub(r'\D', '', p)
    if len(digits) == 10 and digits.startswith('0'):
        return f"{digits[:2]} {digits[2:6]} {digits[6:]}"
    return p

def export_leads(leads: list, path: str):
    if not leads:
        print("No leads to export")
        return pd.DataFrame()

    df = pd.DataFrame(leads)
    
    # --- CLEANING ---
    if "Rating" in df.columns:
        df["Rating"] = df["Rating"].apply(clean_rating)
    if "Website" in df.columns:
        df["Website Clean"] = df["Website"].apply(clean_website)
    else:
        df["Website Clean"] = ""
    if "Phone" in df.columns:
        df["Phone Formatted"] = df["Phone"].apply(format_phone)
    else:
        df["Phone Formatted"] = ""
    
    # Dedup
    if "Business Name" in df.columns and "Address" in df.columns:
        df = df.drop_duplicates(subset=["Business Name","Address"])
    
    # Sort by rating
    if "Rating" in df.columns:
        df = df.sort_values(by="Rating", ascending=False, na_position='last')

    # --- PREPARE PRO DATAFRAME ---
    # Map to professional columns
    pro_data = []
    for _, row in df.iterrows():
        pro_data.append({
            "Business Name": row.get("Business Name",""),
            "Category": row.get("Category",""),
            "Full Address": row.get("Address",""),
            "Phone": row.get("Phone Formatted", row.get("Phone","")),
            "Website": row.get("Website Clean", row.get("Website","")),
            "Rating (5)": row.get("Rating",""),
            "Phone Verified": "Yes" if row.get("Phone") and len(re.sub(r'\D','',str(row.get("Phone"))))>=8 else "No",
            "Source": row.get("Source","Google Maps")
        })
    df_pro = pd.DataFrame(pro_data)

    # --- EXPORT CSV (CLEAN) ---
    try:
        df_pro.to_csv(path, index=False, encoding='utf-8-sig')
        print(f"✅ Exported {len(df_pro)} leads to {path}")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(path)
        path = f"{base}_{timestamp}{ext}"
        df_pro.to_csv(path, index=False, encoding='utf-8-sig')
        print(f"⚠️ File locked, saved as {path}")

    # --- EXPORT PROFESSIONAL EXCEL ---
    try:
        xlsx_path = os.path.splitext(path)[0] + "_PRO.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        thin = Side(border_style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        fill_gray = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        # Header
        for col_idx, col_name in enumerate(df_pro.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

        # Data
        for r_idx, row in enumerate(df_pro.itertuples(index=False), 2):
            fill = fill_gray if r_idx % 2 == 0 else fill_white
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if c_idx == 1:
                    cell.font = Font(bold=True, size=11)
                if c_idx == 6:
                    try:
                        if value and float(value) >= 4.6:
                            cell.font = Font(bold=True, color="16A34A")
                    except:
                        pass

        # Auto width + filter + freeze
        from openpyxl.utils import get_column_letter
        for col_idx, col_name in enumerate(df_pro.columns, 1):
            max_len = max([len(str(col_name))] + [len(str(v)[:50]) if v is not None else 0 for v in df_pro.iloc[:, col_idx-1].astype(str).tolist()])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
        
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'

        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2['A1'] = "GMaps Leads Scraper Pro - Report"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A3'] = f"Total Leads: {len(df_pro)}"
        ws2['A4'] = f"Verified Phones: {len(df_pro[df_pro['Phone Verified']=='Yes'])}"
        ws2['A5'] = f"Average Rating: {df_pro['Rating (5)'].mean():.2f}" if not df_pro['Rating (5)'].isna().all() else "Average Rating: N/A"
        ws2['A6'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        wb.save(xlsx_path)
        print(f"✅ Professional Excel saved to {xlsx_path}")

    except Exception as e:
        print(f"Excel export failed: {e}")

    return df_pro
